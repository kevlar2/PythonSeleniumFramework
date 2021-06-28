import os
import openpyxl


class HomePageData:

    test_homepage_data = [{"fullname": "Kevin Oru",
                           "Email address": "Kevin.Oru@hpdlendscape.com",
                           "Password": "password",
                           "Gender": "Male",
                           "Gender_value": 0,
                           "DOB": "22/04/1984"},
                          {"fullname": "Aryah Oru",
                           "Email address": "Aryah.Oru@hpdlendscape.com",
                           "Password": "Aryah",
                           "Gender": "Female",
                           "Gender_value": 1,
                           "DOB": "01/02/2020"}]

    @staticmethod
    def get_test_data_from_excel(test_case_name):
        data_dict = {}
        path = os.getcwd()[:60]
        print(path)
        book = openpyxl.load_workbook(path + '\\TestData\\Test_data_excel.xlsx')
        sheet = book.active
        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                # data_dict[data_key] = data_value
                for j in range(2, sheet.max_column+1):
                    data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[data_dict]
