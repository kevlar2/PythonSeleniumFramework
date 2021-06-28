import openpyxl


book = openpyxl.load_workbook("Test_data_excel.xlsx")
sheet = book.active
data_dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Kevin"
book.save("Test_data_excel.xlsx")
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet["A5"].value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "TestCase6":

        for j in range(2, sheet.max_column+1):
            data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(data_dict)
